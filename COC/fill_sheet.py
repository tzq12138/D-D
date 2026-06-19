"""Fill COC character sheet template with a generated investigator."""
import random, sys, io, copy
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.utils import get_column_letter

# ============================================================
# Generator (same as gen_investigator.py but returns raw data)
# ============================================================
SKILL_BASE = {
    "会计学":5,"人类学":1,"估价":5,"考古学":1,"取悦":15,"攀爬":20,
    "计算机使用":5,"信用评级":0,"克苏鲁神话":0,"乔装":5,"闪避":"DEX/2","汽车驾驶":20,
    "电气维修":10,"电子学":1,"话术":5,"格斗":25,"射击":20,"急救":30,"历史":5,
    "恐吓":15,"跳跃":20,"外语":1,"母语":"EDU","法律":5,"图书馆使用":20,"聆听":20,
    "锁匠":1,"机械维修":10,"医学":1,"博物学":10,"导航":10,"神秘学":5,"操作重型机械":1,
    "说服":10,"精神分析":1,"心理学":10,"骑术":5,"妙手":10,
    "侦查":25,"潜行":20,"生存":10,"游泳":20,"投掷":20,"追踪":10,
}

def roll(n, d, plus=0):
    return sum(random.randint(1, d) for _ in range(n)) + plus

def gen_character():
    stats = {
        "STR": roll(3,6)*5, "CON": roll(3,6)*5, "SIZ": roll(2,6,6)*5,
        "DEX": roll(3,6)*5, "APP": roll(3,6)*5, "INT": roll(2,6,6)*5,
        "POW": roll(3,6)*5, "EDU": roll(2,6,6)*5, "LUCK": roll(3,6)*5,
    }
    hp = (stats["CON"] + stats["SIZ"]) // 10
    san = stats["POW"]
    mp = stats["POW"] // 5

    total = stats["STR"] + stats["SIZ"]
    if total < 65:   db, build = "-2", -2
    elif total < 85: db, build = "-1", -1
    elif total < 125: db, build = "0", 0
    elif total < 165: db, build = "+1D4", 1
    elif total < 205: db, build = "+1D6", 2
    else:            db, build = "+2D6", 3

    above = sum(1 for v in [stats["STR"], stats["DEX"]] if v >= stats["SIZ"])
    mov = 7 if above == 0 else (8 if above == 1 else 9)

    # Pick occupation — use criminal subtypes from template
    occ_num = 33  # 罪犯-独行罪犯
    occ_name = "罪犯-独行罪犯"
    credit = random.randint(5, 65)

    occ_pts = stats["EDU"] * 2 + stats["DEX"] * 2
    int_pts = stats["INT"] * 2

    # Skills to allocate
    oskills = ["话术","心理学","侦查","潜行","格斗","射击","锁匠","妙手"]

    # Allocate occupational points
    skills = {}
    for sk in oskills:
        skills[sk] = {"base": SKILL_BASE.get(sk, 1), "occ": 0, "int": 0, "total": SKILL_BASE.get(sk, 1)}

    # Split occ_pts across 8 skills roughly
    per_skill = occ_pts // 8
    rem = occ_pts % 8
    alloc = [per_skill + (1 if i < rem else 0) for i in range(8)]
    # Add variance
    for i in range(len(alloc)):
        shift = random.randint(-15, 15)
        if 0 <= i+1 < len(alloc):
            alloc[i] += shift
            alloc[i+1] -= shift
    alloc = [max(0, a) for a in alloc]

    for sk, pts in zip(oskills, alloc):
        base = SKILL_BASE.get(sk, 1)
        if isinstance(base, str): base = 25  # fallback
        skills[sk]["occ"] = pts
        skills[sk]["total"] = base + pts

    # Interest points
    interest_pool = ["聆听","闪避","急救","图书馆使用","汽车驾驶","攀爬",
                     "跳跃","恐吓","估价","博物学","神秘学","历史","投掷","游泳"]
    chosen = random.sample(interest_pool, random.randint(4, 6))
    ipts = int_pts
    for i, c in enumerate(chosen):
        pts = ipts // (len(chosen) - i)
        pts = max(5, pts)
        if c not in skills:
            base = SKILL_BASE.get(c, 5)
            if isinstance(base, str): base = 25
            skills[c] = {"base": base, "occ": 0, "int": 0, "total": base}
        skills[c]["int"] = pts
        skills[c]["total"] = skills[c]["base"] + skills[c]["occ"] + pts
        ipts -= pts

    # Credit rating
    skills["信用评级"] = {"base": 0, "occ": credit, "int": 0, "total": credit}

    # Ensure 闪避 has base = DEX/2
    if "闪避" not in skills:
        skills["闪避"] = {"base": stats["DEX"]//2, "occ": 0, "int": 0, "total": stats["DEX"]//2}
    else:
        skills["闪避"]["base"] = stats["DEX"] // 2
        skills["闪避"]["total"] = skills["闪避"]["base"] + skills["闪避"].get("occ",0) + skills["闪避"].get("int",0)

    # Ensure 母语 has base = EDU
    if "母语" not in skills:
        skills["母语"] = {"base": stats["EDU"], "occ": 0, "int": 0, "total": stats["EDU"]}
    else:
        skills["母语"]["base"] = stats["EDU"]
        skills["母语"]["total"] = skills["母语"]["base"] + skills["母语"].get("occ",0) + skills["母语"].get("int",0)

    return stats, hp, san, mp, db, build, mov, occ_name, occ_num, credit, skills

# ============================================================
# Template fill logic
# ============================================================

# Skill name → (side, row, sub_type_col)
# side: 'L'=left(F col), 'R'=right(AB col)
SKILL_MAP = {}
wb_ref = openpyxl.load_workbook(r'F:\D&D\COC\tzq12138.xlsx')
ws_ref = wb_ref['人物卡']

# Left side
for row in range(16, 50):
    f = ws_ref.cell(row=row, column=6).value
    h = ws_ref.cell(row=row, column=8).value
    if f and not str(f).startswith('='):
        key = f.rstrip('：①②③Ω ').strip()
        SKILL_MAP[key] = ('L', row, h)
# Right side
for row in range(16, 49):
    ab = ws_ref.cell(row=row, column=28).value
    if ab and not str(ab).startswith('='):
        key = ab.rstrip('：①②③Ω ').strip()
        SKILL_MAP[key] = ('R', row, None)

# Custom name mappings for template quirks
SKILL_NAME_MAP = {
    "格斗": "格斗：",
    "射击": "射击：",
    "外语": "外语①",
}

def fill_template(stats, hp, san, mp, db, build, mov, occ_name, occ_num, credit, skills):
    wb = openpyxl.load_workbook(r'F:\D&D\COC\tzq12138.xlsx')
    ws = wb['人物卡']

    # --- Basic info ---
    ws['E3'] = '阿尔伯特·格雷'     # 姓名
    ws['E4'] = '玩家'              # 玩家
    ws['E5'] = occ_name            # 职业
    ws['M5'] = occ_num             # 职业序号
    ws['E6'] = 26                  # 年龄
    ws['M6'] = '男'                # 性别
    ws['E7'] = '阿卡姆'            # 住地
    ws['M7'] = '波士顿'            # 故乡

    # --- Attributes ---
    ws['U3'] = stats['STR']
    ws['U5'] = stats['CON']
    ws['U7'] = stats['SIZ']
    ws['AA3'] = stats['DEX']
    ws['AA5'] = stats['APP']
    ws['AA7'] = stats['INT']
    ws['AG3'] = stats['POW']
    ws['AG5'] = stats['EDU']
    ws['AG7'] = stats['LUCK']

    # --- HP / SAN / MP / MOV ---
    ws['E10'] = hp
    ws['N10'] = san
    ws['W10'] = mp

    # --- Skills ---
    # Helper to safely set cell value (handles merged cells)
    def safe_set(ws, row, col, value):
        cell = ws.cell(row=row, column=col)
        try:
            cell.value = value
        except AttributeError:
            # MergedCell - find the parent
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    parent = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    parent.value = value
                    return
            print(f"  [WARN] Cannot write to merged cell {cell.coordinate}")

    # Fill skill points
    for skill_name, sdata in skills.items():
        if skill_name not in SKILL_MAP:
            print(f"  [WARN] Skill '{skill_name}' not found in template, skipping")
            continue

        side, row, sub_type = SKILL_MAP[skill_name]

        if side == 'L':
            safe_set(ws, row, 14, sdata.get('occ', 0))
            safe_set(ws, row, 16, sdata.get('int', 0))
            if sub_type and skill_name == "格斗":
                safe_set(ws, row, 8, "拳击")
            elif sub_type and skill_name == "射击":
                safe_set(ws, row, 8, "手枪")
        else:
            safe_set(ws, row, 36, sdata.get('occ', 0))
            safe_set(ws, row, 38, sdata.get('int', 0))

    # Save
    out_path = r'F:\D&D\COC\tzq12138.xlsx'
    wb.save(out_path)
    print(f"Saved to {out_path}")
    return True

if __name__ == "__main__":
    random.seed()
    data = gen_character()
    stats, hp, san, mp, db, build, mov, occ_name, occ_num, credit, skills = data

    print("Generated character:")
    print(f"  {occ_name} (职业序号 {occ_num})")
    print(f"  STR={stats['STR']} CON={stats['CON']} SIZ={stats['SIZ']} DEX={stats['DEX']}")
    print(f"  APP={stats['APP']} INT={stats['INT']} POW={stats['POW']} EDU={stats['EDU']} LUCK={stats['LUCK']}")
    print(f"  HP={hp} SAN={san} MP={mp} MOV={mov} DB={db} Build={build}")
    print(f"  Credit={credit}%")
    print(f"\nSkills:")
    for name, sd in sorted(skills.items(), key=lambda x: -x[1]['total']):
        print(f"  {name}: {sd['total']}% (base={sd['base']} occ={sd['occ']} int={sd['int']})")

    ok = fill_template(stats, hp, san, mp, db, build, mov, occ_name, occ_num, credit, skills)
    if ok:
        print("\nDone!")
