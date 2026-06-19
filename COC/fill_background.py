"""Fill background story into character sheet."""
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'F:\D&D\COC\tzq12138.xlsx')
ws = wb['人物卡']

def safe_set(ws, row, col, value):
    try:
        ws.cell(row=row, column=col).value = value
    except AttributeError:
        for mr in ws.merged_cells.ranges:
            if ws.cell(row=row, column=col).coordinate in mr:
                ws.cell(row=mr.min_row, column=mr.min_col).value = value
                return

# AA col = 27

# W61 / AA61: 个人描述/角色外貌
safe_set(ws, 61, 27, "身高六尺三的魁梧大汉，但肤色苍白，常年带着黑眼圈和低咳。"
    "灰蓝色的眼睛里有一种让人不安的沉静。穿着旧但干净的外套，左手虎口有道陈年刀疤。")

# AA63: 思想信念
safe_set(ws, 63, 27, "「规则是写给穷人和傻瓜看的。聪明人只遵守一条：不要被抓到。」")

# AA65: 重要之人
safe_set(ws, 65, 27, "文森特·罗西——地下钱庄老板，格雷的债主兼唯一信得过他办事的人。"
    "两人之间有一种冷酷的尊重，但尊重是有利息的。")

# AA67: 意义非凡之地
safe_set(ws, 67, 27, "密斯卡托尼克河畔第三座桥下的长椅。格雷深夜无处可去时就来这里坐着，"
    "看对岸大学图书馆的灯火。有时候他会想，如果当年有机会念书，一切会不一样——"
    "然后他把这个念头像烟头一样踩灭。")

# AA69: 宝贵之物
safe_set(ws, 69, 27, "一把老旧的撬锁工具，装在一个磨损的皮套里。不是他自己的——"
    "是他十岁时从一个被捕的窃贼那里偷来的，那个窃贼是他父亲。")

# AA71: 特质
safe_set(ws, 71, 27, "沉默寡言但不内向。格雷不说话的时候是在观察——"
    "他在数房间里有几扇门、几扇窗、几个可以藏东西的角落。"
    "这个习惯让他在监狱里活过了三年。")

# AA73: 难言之隐
safe_set(ws, 73, 27, "他父亲被捕那晚本可以跑的——但格雷偷了他的工具，他父亲没了工具打不开手铐。"
    "那年他十岁。此后二十年的所有错误，可能都从那晚开始。")

# AA75: 伤口和疤痕
safe_set(ws, 75, 27, "左手虎口的刀疤（十四岁第一次被搜身时自己割的，为了甩掉赃物）；"
    "右肩胛骨内侧的枪伤疤痕（出狱后的『欢迎礼物』）；"
    "左肺有慢性炎症，阴雨天会咳血——查尔斯顿监狱的冬天留下的纪念品。")

wb.save(r'F:\D&D\COC\tzq12138.xlsx')
print("Background filled successfully!")
